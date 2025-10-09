"""
Excel Writer Module

Модуль для запису проаналізованих даних назад у Excel файли.
Розширює функціонал excel_reader для запису даних.
"""

import openpyxl
import os
import logging
from pathlib import Path
from typing import Dict, Any, Optional
from datetime import datetime
from openpyxl.styles import PatternFill


class ExcelDataWriter:
    """
    Клас для запису даних у Excel файли
    
    Attributes:
        filename (str): Шлях до Excel файлу
        workbook: Об'єкт робочої книги openpyxl
        worksheet: Активний аркуш Excel
        verbose (bool): Режим детального виводу інформації
    """
    
    def __init__(self, filename: str, verbose: bool = False):
        """
        Ініціалізує ExcelDataWriter
        
        Args:
            filename (str): Шлях до Excel файлу
            verbose (bool): Увімкнути детальний вивід інформації
            
        Raises:
            FileNotFoundError: Якщо файл не знайдено
            Exception: Якщо файл не вдається відкрити
        """
        self.filename = filename
        self.verbose = verbose
        self.workbook = None
        self.worksheet = None
        
        self._load_workbook()
    
    def _load_workbook(self) -> None:
        """Завантажує Excel файл"""
        if not Path(self.filename).exists():
            raise FileNotFoundError(f"Файл '{self.filename}' не знайдено!")
        
        try:
            self.workbook = openpyxl.load_workbook(self.filename)
            self.worksheet = self.workbook.active
            if self.verbose:
                print(f"Успішно завантажено файл для запису: {self.filename}")
        except Exception as e:
            raise Exception(f"Помилка при відкритті файлу: {e}")
    
    def write_data_to_row(self, data: Dict[str, Any], target_row: int, 
                         header_row: int = 2, filename: str = None) -> bool:
        """
        Записує проаналізовані дані у вказаний рядок Excel файлу
        
        Args:
            data (Dict[str, Any]): Дані для запису (результат з transcriber.fill_excel_data)
            target_row (int): Номер рядка для запису
            header_row (int): Номер рядка з заголовками для співставлення стовпців
            filename (str): Назва файлу для заповнення поля "Назва файлу"
            
        Returns:
            bool: True якщо запис успішний, False інакше
        """
        try:
            if not self.worksheet:
                logging.error("Робочий аркуш не завантажено")
                return False
            
            # Створюємо мапу заголовків до номерів стовпців
            header_map = {}
            for col in range(1, self.worksheet.max_column + 1):
                header_value = self.worksheet.cell(row=header_row, column=col).value
                if header_value:
                    header_map[str(header_value).strip()] = col
            
            written_fields = 0
            
            # Записуємо дані
            for field_name, field_data in data.items():
                if field_name in header_map:
                    col_num = header_map[field_name]
                    
                    # Перевіряємо чи потрібно пропустити це поле
                    if self._should_skip_field(field_name, target_row, col_num):
                        if self.verbose:
                            print(f"Пропускаємо поле '{field_name}' (містить 'пропускаємо' або формулу)")
                        continue
                    
                    # Отримуємо значення для запису
                    if isinstance(field_data, dict):
                        value_to_write = field_data.get('analyzed_value')
                    else:
                        value_to_write = field_data
                    
                    # Записуємо значення
                    if value_to_write is not None and value_to_write != "":
                        cell = self.worksheet.cell(row=target_row, column=col_num, value=value_to_write)
                        
                        # Застосовуємо умовне форматування
                        self._apply_conditional_formatting(cell, field_name, field_data, data)
                        
                        written_fields += 1
                        
                        if self.verbose:
                            print(f"Записано '{field_name}': {value_to_write} у комірку {target_row},{col_num}")
                else:
                    if self.verbose:
                        print(f"Заголовок '{field_name}' не знайдено в Excel файлі")
            
            # Заповнюємо поле "Назва файлу" якщо воно є
            if filename:
                written_fields = self._write_filename_field(filename, target_row, header_map, written_fields)
            
            # Підраховуємо бали і оновлюємо поле оцінки
            if '_total_score' in data:
                total_score = data['_total_score']['analyzed_value']
                self._update_score_field(total_score, target_row, header_map)
            
            if self.verbose:
                print(f"Записано {written_fields} полів у рядок {target_row}")
            
            return written_fields > 0
            
        except Exception as e:
            logging.error(f"Помилка при записі даних у Excel: {e}")
            return False
    
    def _should_skip_field(self, field_name: str, target_row: int, col_num: int) -> bool:
        """
        Перевіряє чи потрібно пропустити поле
        
        Args:
            field_name (str): Назва поля
            target_row (int): Номер рядка
            col_num (int): Номер стовпця
            
        Returns:
            bool: True якщо поле потрібно пропустити
        """
        # Пропускаємо поле "Оцінка" (зазвичай з формулою)
        if "оцінка" in field_name.lower() or "оценка" in field_name.lower():
            return True
            
        # Перевіряємо чи в комірці є текст "пропускаємо"
        current_value = self.worksheet.cell(row=target_row, column=col_num).value
        if current_value and isinstance(current_value, str):
            if "пропускаємо" in current_value.lower() or "пропускаем" in current_value.lower():
                return True
        
        # Перевіряємо чи в комірці є формула (починається з =)
        if current_value and isinstance(current_value, str) and current_value.startswith('='):
            return True
            
        return False
    
    def _write_filename_field(self, filename: str, target_row: int, header_map: dict, written_fields: int) -> int:
        """
        Записує назву файлу в відповідне поле
        
        Args:
            filename (str): Назва файлу
            target_row (int): Номер рядка
            header_map (dict): Мапа заголовків
            written_fields (int): Кількість вже записаних полів
            
        Returns:
            int: Оновлена кількість записаних полів
        """
        # Шукаємо поле для назви файлу
        filename_fields = ["назва файлу", "название файла", "filename", "file name", "файл"]
        
        for field_candidate in filename_fields:
            for header_name, col_num in header_map.items():
                if field_candidate.lower() in header_name.lower():
                    # Записуємо тільки назву файлу без шляху та розширення
                    clean_filename = os.path.splitext(os.path.basename(filename))[0]
                    self.worksheet.cell(row=target_row, column=col_num, value=clean_filename)
                    written_fields += 1
                    
                    if self.verbose:
                        print(f"Записано назву файлу '{clean_filename}' у поле '{header_name}'")
                    return written_fields
        
        return written_fields
    
    def find_next_empty_row(self, header_row: int = 2) -> int:
        """
        Знаходить наступний порожній рядок для запису
        
        Args:
            header_row (int): Номер рядка з заголовками
            
        Returns:
            int: Номер наступного порожнього рядка
        """
        # Починаємо пошук з рядка після заголовків
        start_row = header_row + 1
        
        for row_num in range(start_row, self.worksheet.max_row + 2):
            # Перевіряємо чи рядок порожній
            is_empty = True
            for col in range(1, self.worksheet.max_column + 1):
                cell_value = self.worksheet.cell(row=row_num, column=col).value
                if cell_value is not None and str(cell_value).strip() != "":
                    is_empty = False
                    break
            
            if is_empty:
                if self.verbose:
                    print(f"Знайдено порожній рядок: {row_num}")
                return row_num
        
        # Якщо всі рядки заповнені, повертаємо наступний
        next_row = self.worksheet.max_row + 1
        if self.verbose:
            print(f"Всі рядки заповнені, створюємо новий: {next_row}")
        return next_row

    def add_new_row_with_data(self, data: Dict[str, Any], header_row: int = 2, filename: str = None) -> Optional[int]:
        """
        Додає новий рядок з даними у наступне порожнє місце таблиці
        
        Args:
            data (Dict[str, Any]): Дані для запису
            header_row (int): Номер рядка з заголовками
            filename (str): Назва файлу для запису
            
        Returns:
            Optional[int]: Номер створеного рядка або None при помилці
        """
        try:
            # Знаходимо наступний порожній рядок
            new_row = self.find_next_empty_row(header_row)
            
            # Записуємо дані у знайдений рядок
            if self.write_data_to_row(data, new_row, header_row, filename):
                if self.verbose:
                    print(f"Заповнено рядок {new_row} з даними")
                return new_row
            else:
                logging.error("Не вдалося записати дані у рядок")
                return None
                
        except Exception as e:
            logging.error(f"Помилка при створенні нового рядка: {e}")
            return None
    
    def add_metadata_column(self, column_name: str = "Дата обробки", 
                           header_row: int = 2) -> bool:
        """
        Додає стовпець з метаданими (наприклад, дата обробки)
        
        Args:
            column_name (str): Назва стовпця
            header_row (int): Номер рядка з заголовками
            
        Returns:
            bool: True якщо стовпець додано успішно
        """
        try:
            # Знаходимо першу порожню колонку
            last_col = self.worksheet.max_column
            new_col = last_col + 1
            
            # Перевіряємо чи останній стовпець дійсно заповнений
            if not self.worksheet.cell(row=header_row, column=last_col).value:
                new_col = last_col
            
            # Додаємо заголовок
            self.worksheet.cell(row=header_row, column=new_col, value=column_name)
            
            if self.verbose:
                print(f"Додано стовпець '{column_name}' у позицію {new_col}")
            
            return True
            
        except Exception as e:
            logging.error(f"Помилка при додаванні стовпця: {e}")
            return False
    
    def save_file(self, new_filename: Optional[str] = None) -> bool:
        """
        Зберігає Excel файл
        
        Args:
            new_filename (Optional[str]): Новий шлях для збереження. 
                                        Якщо None, перезаписує оригінальний файл.
            
        Returns:
            bool: True якщо збереження успішне
        """
        try:
            save_path = new_filename if new_filename else self.filename
            
            # Створюємо директорію якщо не існує
            os.makedirs(os.path.dirname(save_path), exist_ok=True)
            
            self.workbook.save(save_path)
            
            if self.verbose:
                print(f"Excel файл збережено: {save_path}")
            
            return True
            
        except Exception as e:
            logging.error(f"Помилка при збереженні Excel файлу: {e}")
            return False
    
    def close(self) -> None:
        """Закриває робочу книгу"""
        if self.workbook:
            self.workbook.close()
            if self.verbose:
                print("Робочу книгу закрито")
    
    def __enter__(self):
        """Підтримка context manager"""
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        """Підтримка context manager"""
        self.close()
    
    def _apply_conditional_formatting(self, cell, field_name: str, field_data: Dict[str, Any], all_data: Dict[str, Any]) -> None:
        """
        Застосовує умовне форматування до комірки
        
        Args:
            cell: Комірка Excel для форматування
            field_name (str): Назва поля
            field_data (Dict[str, Any]): Дані поля
            all_data (Dict[str, Any]): Всі дані для аналізу
        """
        try:
            # Перевіряємо чи є оцінка менеджера в даних
            manager_evaluation = all_data.get('_manager_evaluation', {}).get('analyzed_value', {})
            
            if isinstance(manager_evaluation, dict) and not manager_evaluation.get('is_performance_good', True):
                # Якщо менеджер працював погано, виділяємо поле коментарів червоним
                comment_keywords = ['коментар', 'коментарі', 'comment', 'comments', 'примітка', 'заувага']
                
                if any(keyword in field_name.lower() for keyword in comment_keywords):
                    # Червоне виділення для полів коментарів
                    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                    cell.fill = red_fill
                    
                    if self.verbose:
                        print(f"Застосовано червоне виділення до поля '{field_name}' через погану оцінку менеджера")
            
            # Додаткове форматування для поля оцінки
            score_keywords = ['оцінка', 'оценка', 'score', 'rating', 'бал']
            if any(keyword in field_name.lower() for keyword in score_keywords):
                # Якщо це підрахований бал, додаємо зелене виділення
                if field_data.get('calculated_score', False):
                    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
                    cell.fill = green_fill
                    
                    if self.verbose:
                        print(f"Застосовано зелене виділення до автоматично підрахованого поля '{field_name}'")
                        
        except Exception as e:
            logging.error(f"Помилка при застосуванні форматування до поля {field_name}: {e}")
    
    def _update_score_field(self, total_score: int, target_row: int, header_map: dict) -> bool:
        """
        Оновлює поле "Оцінка" підрахованим балом
        
        Args:
            total_score (int): Підрахований загальний бал
            target_row (int): Номер рядка
            header_map (dict): Мапа заголовків
            
        Returns:
            bool: True якщо поле було оновлено
        """
        score_keywords = ['оцінка', 'оценка', 'score', 'rating', 'бал']
        
        for header_name, col_num in header_map.items():
            if any(keyword in header_name.lower() for keyword in score_keywords):
                # Перевіряємо чи це не поле що має пропускатися
                if not self._should_skip_field(header_name, target_row, col_num):
                    cell = self.worksheet.cell(row=target_row, column=col_num, value=str(total_score))
                    
                    # Зелене виділення для автоматично підрахованих балів
                    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
                    cell.fill = green_fill
                    
                    if self.verbose:
                        print(f"Автоматично оновлено поле '{header_name}' балом: {total_score}")
                    return True
        
        return False


def write_analyzed_data_to_excel(excel_file_path: str, analyzed_data: Dict[str, Any], 
                                output_file_path: Optional[str] = None, 
                                target_row: Optional[int] = None,
                                header_row: int = 2,
                                filename: str = None,
                                verbose: bool = False) -> Dict[str, Any]:
    """
    Зручна функція для запису проаналізованих даних у Excel файл
    
    Args:
        excel_file_path (str): Шлях до вихідного Excel файлу
        analyzed_data (Dict[str, Any]): Проаналізовані дані від transcriber.fill_excel_data
        output_file_path (Optional[str]): Шлях для збереження. Якщо None, перезаписує оригінал
        target_row (Optional[int]): Рядок для запису. Якщо None, знаходить наступний порожній
        header_row (int): Рядок з заголовками
        filename (str): Назва файлу для заповнення поля "Назва файлу"
        verbose (bool): Детальний вивід
        
    Returns:
        Dict[str, Any]: Результат операції
    """
    result = {
        'success': False,
        'written_row': None,
        'written_fields': 0,
        'output_file': None,
        'error': None
    }
    
    try:
        with ExcelDataWriter(excel_file_path, verbose=verbose) as writer:
            
            if target_row:
                # Записуємо у вказаний рядок
                success = writer.write_data_to_row(analyzed_data, target_row, header_row, filename)
                if success:
                    result['written_row'] = target_row
            else:
                # Знаходимо наступний порожній рядок
                new_row = writer.add_new_row_with_data(analyzed_data, header_row, filename)
                if new_row:
                    result['written_row'] = new_row
                    success = True
                else:
                    success = False
            
            if success:
                # Зберігаємо файл (завжди в оригінал згідно з вимогами)
                save_path = excel_file_path  # Завжди перезаписуємо оригінальний файл
                if writer.save_file(save_path):
                    result['success'] = True
                    result['output_file'] = save_path
                    result['written_fields'] = len(analyzed_data)
                    
                    if verbose:
                        print(f"✅ Дані успішно записано в оригінальну таблицю: {save_path}")
                        print(f"   📍 Заповнено рядок: {result['written_row']}")
                        if filename:
                            print(f"   📁 Файл: {os.path.basename(filename)}")
                else:
                    result['error'] = "Помилка при збереженні файлу"
            else:
                result['error'] = "Помилка при записі даних"
        
    except Exception as e:
        result['error'] = str(e)
        logging.error(f"Помилка при записі у Excel: {e}")
    
    return result


if __name__ == "__main__":
    # Приклад використання
    print("=== Тестування Excel Writer ===")
    
    # Приклад даних (як би їх повернув transcriber.fill_excel_data)
    sample_data = {
        "Ім'я": {
            'analyzed_value': 'Іван Петренко',
            'field_type': 'text'
        },
        "Телефон": {
            'analyzed_value': '+380501234567', 
            'field_type': 'text'
        },
        "Статус": {
            'analyzed_value': 'Активний',
            'field_type': 'dropdown'
        }
    }
    
    excel_file = "google_folder/Звіт прослуханих розмов.xlsx"
    
    if os.path.exists(excel_file):
        result = write_analyzed_data_to_excel(
            excel_file_path=excel_file,
            analyzed_data=sample_data,
            output_file_path="output/updated_report.xlsx",
            verbose=True
        )
        
        print(f"Результат: {result}")
    else:
        print(f"Файл {excel_file} не знайдено для тестування")