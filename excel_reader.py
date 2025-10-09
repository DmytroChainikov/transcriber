"""
Excel Data Reader Module

Модуль для читання даних з Excel файлів з підтримкою випадаючих списків.
Використовує openpyxl для обробки .xlsx файлів.

Usage:
    from excel_reader import ExcelDataReader
    
    reader = ExcelDataReader("file.xlsx")
    data = reader.read_data(header_row=2, data_row=3)
"""

import openpyxl
from pathlib import Path
from openpyxl.worksheet.datavalidation import DataValidation
from typing import Dict, List, Optional, Union, Any


class ExcelDataReader:
    """
    Клас для читання та обробки даних з Excel файлів
    
    Attributes:
        filename (str): Шлях до Excel файлу
        workbook: Об'єкт робочої книги openpyxl
        worksheet: Активний аркуш Excel
        verbose (bool): Режим детального виводу інформації
    """
    
    def __init__(self, filename: str, verbose: bool = False):
        """
        Ініціалізує ExcelDataReader
        
        Args:
            filename (str): Шлях до Excel файлу
            verbose (bool): Увімкнути детальний вивід інформації
            
        Raises:
            FileNotFoundError: Якщо файл не знайдено
            Exception: Якщо файл не вдається відкрити
        """
        self.filename = filename
        self.verbose = verbose
        self.workbook = None
        self.worksheet = None
        
        self._load_workbook()
    
    def _load_workbook(self) -> None:
        """Завантажує Excel файл"""
        if not Path(self.filename).exists():
            raise FileNotFoundError(f"Файл '{self.filename}' не знайдено!")
        
        try:
            self.workbook = openpyxl.load_workbook(self.filename)
            self.worksheet = self.workbook.active
            if self.verbose:
                print(f"Успішно завантажено файл: {self.filename}")
        except Exception as e:
            raise Exception(f"Помилка при відкритті файлу: {e}")
    
    def get_dropdown_values(self, row: int, col: int) -> Optional[List[str]]:
        """
        Отримує значення випадаючого списку для конкретної комірки
        
        Args:
            row (int): Номер рядка
            col (int): Номер стовпця
        
        Returns:
            Optional[List[str]]: Список доступних значень випадаючого списку або None
        """
        if not self.worksheet:
            return None
            
        cell_coord = self.worksheet.cell(row=row, column=col).coordinate
        
        # Перевіряємо всі правила валідації даних на аркуші
        for validation in self.worksheet.data_validations.dataValidation:
            # Перевіряємо чи комірка входить до діапазону валідації
            if cell_coord in validation.cells:
                # Якщо це список значень
                if validation.type == "list" and validation.formula1:
                    formula = validation.formula1
                    # Видаляємо лапки та розділяємо по комах
                    if formula.startswith('"') and formula.endswith('"'):
                        values = formula[1:-1].split(',')
                        return [v.strip() for v in values]
                    # Якщо формула посилається на діапазон
                    elif ":" in formula:
                        try:
                            # Отримуємо діапазон комірок
                            range_cells = self.worksheet[formula]
                            if hasattr(range_cells, '__iter__'):
                                values = []
                                for cell_row in range_cells:
                                    if hasattr(cell_row, '__iter__'):
                                        for cell in cell_row:
                                            if cell.value is not None:
                                                values.append(str(cell.value).strip())
                                    else:
                                        if cell_row.value is not None:
                                            values.append(str(cell_row.value).strip())
                                return values
                            else:
                                return [str(range_cells.value).strip()] if range_cells.value else []
                        except Exception:
                            if self.verbose:
                                print(f"Помилка при обробці формули діапазону: {formula}")
        return None
    
    def read_data(self, header_row: int = 2, data_row: int = 3) -> Dict[str, Any]:
        """
        Читає дані з Excel файлу та створює словник з парами ключ-значення
        
        Args:
            header_row (int): Номер рядка з заголовками (за замовчуванням 2)
            data_row (int): Номер рядка з даними (за замовчуванням 3)
        
        Returns:
            Dict[str, Any]: Словник з ключами з header_row та значеннями з data_row
        """
        if not self.worksheet:
            return {}
        
        header_values = []
        data_values = []
        
        # Проходимо по всіх стовпцях до тих пір, поки не зустрінемо порожню комірку в заголовку
        for col in range(1, self.worksheet.max_column + 1):
            header_cell_value = self.worksheet.cell(row=header_row, column=col).value
            if header_cell_value is not None:
                header_values.append(str(header_cell_value).strip())
                
                # Читаємо відповідне значення з data_row
                data_cell_value = self.worksheet.cell(row=data_row, column=col).value
                
                # Перевіряємо чи є випадаючий список для цієї комірки
                dropdown_values = self.get_dropdown_values(data_row, col)
                
                if data_cell_value is not None:
                    # Перевіряємо чи це текст, число чи інше значення
                    if isinstance(data_cell_value, str):
                        processed_value = data_cell_value.strip()
                    else:
                        processed_value = data_cell_value
                    
                    # Якщо є випадаючий список, додаємо інформацію про доступні варіанти
                    if dropdown_values:
                        data_values.append({
                            'value': processed_value,
                            'dropdown_options': dropdown_values,
                            'type': 'dropdown'
                        })
                    else:
                        data_values.append({
                            'value': processed_value,
                            'type': 'text'
                        })
                else:
                    # Якщо комірка порожня, але є випадаючий список
                    if dropdown_values:
                        data_values.append({
                            'value': None,
                            'dropdown_options': dropdown_values,
                            'type': 'dropdown'
                        })
                    else:
                        data_values.append({
                            'value': None,
                            'type': 'empty'
                        })
            else:
                break
        
        # Створюємо словник з парами ключ-значення
        # Обробляємо дублікати в ключах, додаючи номер до повторюваних ключів
        keys_dict = {}
        key_counts = {}
        
        for i, key in enumerate(header_values):
            if key in key_counts:
                key_counts[key] += 1
                unique_key = f"{key}_{key_counts[key]}"
            else:
                key_counts[key] = 1
                unique_key = key
            
            # Присвоюємо значення з data_row
            if i < len(data_values):
                keys_dict[unique_key] = data_values[i]
            else:
                keys_dict[unique_key] = None
        
        if self.verbose:
            self._print_results(header_values, data_values)
        
        return keys_dict
    
    def _print_results(self, header_values: List[str], data_values: List[Dict]) -> None:
        """Виводить результати читання (тільки у verbose режимі)"""
        print(f"Знайдено {len(header_values)} ключів у заголовках:")
        for i, key in enumerate(header_values, 1):
            print(f"{i}. {key}")
        
        print(f"\nЗначення з рядка даних:")
        for i, value in enumerate(data_values, 1):
            if isinstance(value, dict):
                if value['type'] == 'dropdown':
                    dropdown_str = ', '.join(value['dropdown_options']) if value['dropdown_options'] else 'Немає варіантів'
                    print(f"{i}. Значення: {value['value']} | Тип: Випадаючий список | Варіанти: [{dropdown_str}]")
                else:
                    print(f"{i}. Значення: {value['value']} | Тип: {value['type']}")
            else:
                print(f"{i}. {value}")
    
    def get_all_sheets(self) -> List[str]:
        """
        Повертає список всіх аркушів у робочій книзі
        
        Returns:
            List[str]: Список назв аркушів
        """
        if not self.workbook:
            return []
        return self.workbook.sheetnames
    
    def set_active_sheet(self, sheet_name: str) -> bool:
        """
        Встановлює активний аркуш
        
        Args:
            sheet_name (str): Назва аркуша
            
        Returns:
            bool: True якщо аркуш встановлено успішно, False інакше
        """
        if not self.workbook:
            return False
            
        try:
            self.worksheet = self.workbook[sheet_name]
            if self.verbose:
                print(f"Встановлено активний аркуш: {sheet_name}")
            return True
        except KeyError:
            if self.verbose:
                print(f"Аркуш '{sheet_name}' не знайдено")
            return False
    
    def close(self) -> None:
        """Закриває робочу книгу"""
        if self.workbook:
            self.workbook.close()
            if self.verbose:
                print("Робочу книгу закрито")
    
    def __enter__(self):
        """Підтримка context manager"""
        return self
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        """Підтримка context manager"""
        self.close()


# Для зворотної сумісності з оригінальним кодом
def read_data_from_excel(filename: str, header_row: int = 2, data_row: int = 3, verbose: bool = False) -> Dict[str, Any]:
    """
    Функція для зворотної сумісності з оригінальним кодом
    
    Args:
        filename (str): Шлях до Excel файлу
        header_row (int): Номер рядка з заголовками
        data_row (int): Номер рядка з даними
        verbose (bool): Детальний вивід
        
    Returns:
        Dict[str, Any]: Словник з даними
    """
    try:
        with ExcelDataReader(filename, verbose=verbose) as reader:
            return reader.read_data(header_row, data_row)
    except Exception as e:
        if verbose:
            print(f"Помилка при обробці файлу: {e}")
        return {}


if __name__ == "__main__":
    # Приклад використання
    import pprint
    
    # Використання класу
    try:
        with ExcelDataReader("google_folder/Звіт прослуханих розмов.xlsx", verbose=True) as reader:
            data = reader.read_data()
            print("\n" + "="*50)
            print("РЕЗУЛЬТАТ:")
            print("="*50)
            pprint.pprint(data)
    except Exception as e:
        print(f"Помилка: {e}")